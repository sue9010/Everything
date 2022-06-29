from openpyxl import Workbook
import openpyxl

wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active
l_row = ws.max_row

ws["A"+str(l_row+1)].value = int(l_row+1)

wb.save('test.xlsx')
wb.close()
