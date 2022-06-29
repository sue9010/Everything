
from openpyxl import *


wb = load_workbook("업체정보.xlsx")
ws = wb.active

l_row = ws.max_row
i = 1

comp_name = input("업체명을 입력해주세요")


for i in range(int(l_row)):
    if i == 0:
        pass
    else:
        comp_name = comp_name.lower()
        if comp_name in ws["A"+str(i)].value.lower():
            print(ws["A"+str(i)].value)
            i += 1
        else:
            pass


wb.close()
