import os
from openpyxl import *
import openpyxl
import tkinter
from tkinter import filedialog  
from tkinter import *

# tkinter로 창 실행
window = tkinter.Tk()

# 창 타이틀 설정
window.title("엑셀로 PDF 파일명 변경")
# 창 크기 설정
window.geometry("530x55")
# 창 리사이즈 불가능 설정
window.resizable(False,False)

def excel_rename():
    # 엑셀 파일 선택 창 열기
    window.dirName = filedialog.askopenfilename()
    
    excel_file_name = window.dirName.split("/")
    print(excel_file_name[-1])

    wb = openpyxl.load_workbook(filename = window.dirName)
    ws = wb.active
    last_row = ws.max_row

    for i in range (2,last_row+1):

        n_fileName=ws.cell(row=i, column=20)

        d_num = str(ws.cell(row=i, column=1).value)
        buyer = str(ws.cell(row=i, column=8).value)
        d_date = str(ws.cell(row=i, column=13).value)
        price = str(ws.cell(row=i, column=19).value)

        n ="수출신고필증_"+ d_date +"_"+d_num+"_"+buyer+"_"+price+"USD.pdf"
        n_fileName.value = n
        # print(n_fileName.value)
        
    wb.save(filename=window.dirName)
    wb.close()

    print(window.dirName)
    lbl.configure(text = window.dirName)

def excel_rename2():
    # 엑셀 파일 선택 창 열기
    window.dirName = filedialog.askopenfilename()
    
    # excel_file_name = window.dirName.split("/")
    # print(excel_file_name[-1])

    wb = openpyxl.load_workbook(filename = window.dirName)
    ws = wb.active
    last_row = ws.max_row

    for i in range (2,last_row+1):

        n_fileName=ws.cell(row=i, column=23)

        d_num = str(ws.cell(row=i, column=1).value)
        r_d_num = d_num.replace("-","")
        buyer = str(ws.cell(row=i, column=8).value)
        d_date = str(ws.cell(row=i, column=4).value)
        r_d_date = d_date.replace("-","")
        price = str(ws.cell(row=i, column=16).value)
        print(r_d_num, r_d_date, price, buyer)
        n ="수입신고필증_"+ r_d_date +"_"+r_d_num+"_"+buyer+"_"+price+"USD.pdf"
        n_fileName.value = n
        print(n_fileName.value)
        # print(n)
       
    wb.save(filename=window.dirName)
    wb.close()

    # print(window.dirName)
    lbl.configure(text = window.dirName)

def pdf_rename():
    window.dirName = filedialog.askdirectory()
    # # 폴더 경로를 받아오기
    file_path = window.dirName
    # 폴더 안에 있는 파일명을 리스트 형태로 받아오기
    file_names = os.listdir(file_path)
    # print(file_names)
    for name in file_names:
        # 폴더 경로 + 파일명 -> 기존 파일명
        old_name = file_path + "/"+ name
        # 폴더 경로 + 슬라이싱 한 파일명 -> 새로운 파일명
        new_name = file_path + "/"+ name[4:18]+".pdf"
        # 파일명 변경 명령문
        os.rename(old_name, new_name)

    wb = openpyxl.load_workbook(filename = lbl.cget("text"))
    ws = wb.active 
    last_row = ws.max_row

    file_list = window.dirName
    file_names = os.listdir(file_list)
    for file_name in file_names:
        f = file_name.split(".")[0]
        for i in range(2,last_row+1):
            old_name = file_path + "/" + f + ".pdf"
            n_fileName = ws.cell(row=i, column=20).value
            renamed = window.dirName + "/" + n_fileName
            d_num = ws.cell(row=i, column=1).value
            if f == d_num:
                os.rename(old_name, renamed)

    wb.close()
    lbl.configure(text = "")

def pdf_rename2():
    window.dirName = filedialog.askdirectory()
    file_path = window.dirName
    file_names = os.listdir(file_path)

    for name in file_names:
        old_name = file_path + "/"+ name
        n = name.split("_")
        new_name = file_path + "\\"+n[1]+".pdf"
        os.rename(old_name, new_name)

    wb = openpyxl.load_workbook(filename = lbl.cget("text"))
    ws = wb.active 
    last_row = ws.max_row

    file_list = window.dirName
    file_names = os.listdir(file_list)
    for file_name in file_names:
        f = file_name.split(".")[0]
        for i in range(2,last_row+1):
            old_name = file_path + "/" + f + ".pdf"
            n_fileName = ws.cell(row=i, column=23).value
            renamed = window.dirName + "/" + n_fileName
            d_num = ws.cell(row=i, column=1).value
            d_d_num = d_num.replace("-","")
            if f == d_d_num:
                os.rename(old_name, renamed)
    wb.close()
    lbl.configure(text = "")

lbl = tkinter.Label(window, text="excel 파일 경로",width = 37, height =1, wraplength=310)
lbl.grid(row=0, column=0)

btn = tkinter.Button(window, text="excel 파일 선택(수출)",width = 17, height =1, command=excel_rename)
btn.grid(row=0, column=1)
btn3 = tkinter.Button(window, text="excel 파일 선택(수입)",width = 17, height =1, command=excel_rename2)
btn3.grid(row=1, column=1)
btn2 = tkinter.Button(window, text ="pdf 폴더 선택(수출)",width = 17, height =1, command=pdf_rename)
btn2.grid(row=0, column=2)
btn4 = tkinter.Button(window, text ="pdf 폴더 선택(수입)",width = 17, height =1, command=pdf_rename2)
btn4.grid(row=1, column=2)

# 창 유지
window.mainloop()

