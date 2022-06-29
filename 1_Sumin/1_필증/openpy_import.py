import openpyxl
from openpyxl import *
import os

folder_path = r"C:\Users\Sue\Desktop\2022 수입신고필증\원본\202204"
file_path = r"C:\Users\Sue\Desktop\2022 수입신고필증\원본\202204.xlsx"

wb = openpyxl.load_workbook(filename = file_path)
ws = wb.active
last_row = ws.max_row

for i in range (2,last_row+1):

    n_fileName=ws.cell(row=i, column=23)

    d_num = str(ws.cell(row=i, column=1).value)
    r_d_num = d_num.replace("-","")
    buyer = str(ws.cell(row=i, column=8).value)
    d_date = str(ws.cell(row=i, column=4).value)
    r_d_date = d_date.replace("-","")
    price = str(ws.cell(row=i, column=16).value)
    # print(r_d_num, r_d_date, price, buyer)
    n ="수입신고필증_"+ r_d_date +"_"+r_d_num+"_"+buyer+"_"+price+"USD.pdf"
    n_fileName.value = n
    print(n_fileName.value)


