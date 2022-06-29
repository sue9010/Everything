import os
from openpyxl import *
import openpyxl
import tkinter
from tkinter import filedialog  
from tkinter import *

# tkinter로 창 실행
window = tkinter.Tk()

# 창 타이틀 설정
window.title("엑셀로 PDF 파일명 변경")
# 창 크기 설정
window.geometry("530x35")
# 창 리사이즈 불가능 설정
window.resizable(False,False)

def excel_export():
    # 엑셀 파일의 경로
    window.dirName = filedialog.askopenfilename()
    
    wb = openpyxl.load_workbook(filename=window.dirName)
    ws = wb.active
    
    for i in range(2,ws.max_row+1):
        n_fileName = ws.cell(row=i, column=20)

        d_num = str(ws.cell(row=i, column=1).value)
        buyer = str(ws.cell(row=i, column=8).value)
        d_date = str(ws.cell(row=i, column=13).value)
        price = str(ws.cell(row=i, column=19).value)

        n ="수출신고필증_"+ d_date +"_"+d_num+"_"+buyer+"_"+price+"USD.pdf"
        n_fileName.value = n

    wb.save(filename=window.dirName)
    wb.close()

    lbl.configure(text = window.dirName)

    # pdf 파일의 경로
    pdf_path = window.dirName[0:-5]
    file_names = os.listdir(pdf_path)

    for name in file_names:
        old_name = pdf_path + "/"+ name
        new_name = pdf_path + "/"+ name[4:18]+".pdf"
        os.rename(old_name,new_name)


    wb = openpyxl.load_workbook(filename = lbl.cget("text"))
    ws = wb.active 

    file_list = pdf_path
    file_names = os.listdir(file_list)

    for file_name in file_names:
        f = file_name.split(".")[0]
        for i in range(2,ws.max_row+1):
            old_name = pdf_path + "/" + f + ".pdf"
            n_fileName = ws.cell(row=i, column=20).value
            renamed = pdf_path + "/" + n_fileName
            d_num = ws.cell(row=i, column=1).value
            if f == d_num:
                os.rename(old_name, renamed)

    wb.close()
    lbl.configure(text = "완료")


def excel_import():
    # 엑셀 파일의 경로
    window.dirName = filedialog.askopenfilename()
    
    wb = openpyxl.load_workbook(filename=window.dirName)
    ws = wb.active
    
    for i in range(2,ws.max_row+1):
        n_fileName = ws.cell(row=i, column=23)

        d_num = str(ws.cell(row=i, column=1).value)
        r_d_num = d_num.replace("-","")
        buyer = str(ws.cell(row=i, column=8).value)
        d_date = str(ws.cell(row=i, column=4).value)
        r_d_date = d_date.replace("-","")
        price = str(ws.cell(row=i, column=16).value)

        n ="수입신고필증_"+ r_d_date +"_"+r_d_num+"_"+buyer+"_"+price+"USD.pdf"
        n_fileName.value = n

    wb.save(filename=window.dirName)
    wb.close()

    lbl.configure(text = window.dirName)

    # pdf 파일의 경로
    pdf_path = window.dirName[0:-5]
    file_names = os.listdir(pdf_path)

    for name in file_names:
        old_name = pdf_path + "/"+ name
        n = name.split("_")
        new_name = pdf_path + "/"+name.split("_")[1]+".pdf"
        os.rename(old_name, new_name)


    wb = openpyxl.load_workbook(filename = lbl.cget("text"))
    ws = wb.active 

    file_list = pdf_path
    file_names = os.listdir(file_list)

    for file_name in file_names:
        f = file_name.split(".")[0]
        for i in range(2,ws.max_row+1):
            old_name = pdf_path + "/" + f + ".pdf"
            n_fileName = ws.cell(row=i, column=23).value
            renamed = pdf_path + "/" + n_fileName
            d_num = ws.cell(row=i, column=1).value
            d_d_num = d_num.replace("-","")
            if f == d_d_num:
                os.rename(old_name, renamed)

    wb.close()
    lbl.configure(text = "완료")



lbl = tkinter.Label(window, text="excel 파일 경로",width = 37, height =1, wraplength=310)
lbl.grid(row=0, column=0)
btn = tkinter.Button(window, text="수출 파일 선택",width = 17, height =1, command=excel_export)
btn.grid(row=0, column=1)
btn3 = tkinter.Button(window, text="수입 파일 선택",width = 17, height =1, command=excel_import)
btn3.grid(row=0, column=2)


# 창 유지
window.mainloop()

