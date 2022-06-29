import openpyxl
import os

wb = openpyxl.load_workbook(filename = r"C:\Users\Sue\Desktop\2022 수출신고필증\원본\202204.xlsx")
ws = wb.active 
last_row = ws.max_row

file_list = os.listdir(r"C:\Users\Sue\Desktop\2022 수출신고필증\원본\202204")
for file in file_list:
    f = file.split('.')[0]
    # print(f)
    for i in range(2,last_row+1):
        new_filename = ws.cell(row=i, column=20).value
        d_num = ws.cell(row=i, column=1).value
        # print(d_num)
        if f == d_num:
            print(f +" is "+new_filename)


