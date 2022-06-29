from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1셀의 정보를 출력
print(ws["A1"].value) # A1셀의 값을 출력
print(ws["A10"].value) # 값이 없을 땐 'None'을 출력

# row = 1,2,3
# column = A,B,C
print(ws.cell(row=1, column=1).value) # ws["A1"].value
print(ws.cell(row=1, column=2).value) # ws["B1"].value

wb.save("sample.xlsx")
wb.close()